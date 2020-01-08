# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook
def make_xlsx(config):
    '''生产表格'''

    wb = Workbook()    #创建文件对象
    ws = wb.active     #获取第一个sheet

    ws['A1'] = '主机名'
    ws['B1'] = '检查项'
    ws['C1'] = '状态'
    ws['D1'] = '值'

    row = 2

    for i in config:
        ws['A'+ str(row)] = i[0]
        ws['B'+ str(row)] = '设备ipv4:'
        ws['C'+ str(row)] = '正常'
        ws['D'+ str(row)] = i[1].split(';')[0]
        row += 1
        ws['A'+ str(row)] = i[0]
        ws['B'+ str(row)] = '设备ipv6:'
        ws['C'+ str(row)] = '正常'
        ws['D'+ str(row)] = i[1].split(';')[1]
        row += 1
        ws['A'+ str(row)] = i[0]
        ws['B'+ str(row)] = '与radius交互的ip:'
        ws['C'+ str(row)] = '正常'
        ws['D'+ str(row)] = i[2]
        row += 1
        ws['A'+ str(row)] = i[0]
        ws['B'+ str(row)] = '上行port编号及描述:'
        if '缺少描述，请确认' in i[3]:
            ws['C'+ str(row)] = '需确认'
        else:
            ws['C'+ str(row)] = '正常'
        ws['D'+ str(row)] = i[3]
        row += 1
        ws['A'+ str(row)] = i[0]
        ws['B'+ str(row)] = 'Radius服务器配置合规性:'
        if '没有' in i[4]:
            ws['C'+ str(row)] = '异常'
        else:
            ws['C'+ str(row)] = '正常'
        ws['D'+ str(row)] = i[4]
        row += 1
        ws['A'+ str(row)] = i[0]
        ws['B'+ str(row)] = 'video配置合规性:'
        if '配置错误' in i[5] or '没有' in i[5] or '缺少' in i[5]:
            ws['C'+ str(row)] = '异常'
        else:
            ws['C'+ str(row)] = '正常'
        ws['D'+ str(row)] = i[5]
        row += 1
        ws['A'+ str(row)] = i[0]
        ws['B'+ str(row)] = '与ies3000接口业务对应的HGU管理配置和IMS配置检查:'
        if '没有' in i[6]:
            ws['C'+ str(row)] = '异常'
        elif '请确认' in i[6]:
            ws['C'+ str(row)] = '需确认'
        else:
            ws['C'+ str(row)] = '正常'
        ws['D'+ str(row)] = i[6]
        row += 1
        # ws['A'+ str(row)] = i[0]
        # ws['B'+ str(row)] = 'Dns检查:'
        state = '正常'
        for j in i[7].split('\n'):
            if j == '\n':
                continue
            if '正常' not in j:
                state = '异常'
            ws['A'+ str(row)] = i[0]
            ws['B'+ str(row)] = 'Dns检查:'
            ws['C'+ str(row)] = state
            ws['D'+ str(row)] = j
            if j != i[7].split('\n')[-1]:
                row += 1
        ws['A'+ str(row)] = i[0]
        ws['B'+ str(row)] = 'telnet acl白名单检查:'
        if '缺少' in i[8] or '没有' in i[8]:
            ws['C'+ str(row)] = '异常'
        else:
            ws['C'+ str(row)] = '正常'
        ws['D'+ str(row)] = i[8]
        row += 1
        # ws['A'+ str(row)] = i[0]
        # ws['B'+ str(row)] = 'PPPoE业务本地默认速率检查:'
        state = '需确认'
        for j in i[9].split('\n'):
            ws['A'+ str(row)] = i[0]
            ws['B'+ str(row)] = 'PPPoE业务本地默认速率检查:'
            ws['C'+ str(row)] = state
            ws['D'+ str(row)] = j
            if j != i[9].split('\n')[-1]:
                row += 1

        # row += 1
        ws['A'+ str(row)] = i[0]
        ws['B'+ str(row)] = 'urpf配置检查:'
        if '缺少' in i[10] or '存在' in i[10]:
            ws['C'+ str(row)] = '异常'
        else:
            ws['C'+ str(row)] = '正常'
        ws['D'+ str(row)] = i[10]

        row += 2

    wb.save('output.xlsx')